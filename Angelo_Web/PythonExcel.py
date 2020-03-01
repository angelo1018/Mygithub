# import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime

#建工作簿
wb = Workbook()
ws1 = wb.create_sheet('hello', 0)
ws = wb['Sheet']

print(ws1)
print(ws)
ws['A1'] = 43
ws.append([1,2,3,4])
ws['A2'] = datetime.datetime.now()

wb.save('sample.xlsx')

# 打开已存在工作簿
wb = load_workbook('sample.xlsx')
#建新表，改表名，改表名格式
wb.create_sheet("Mysheet1", 0) # insert at the penultimate position
ws = wb['hello']
ws.title = 'New Title'
ws.sheet_properties.tabColor = "1072BA"
wb.save('sample1.xlsx')
# 以列表形式展示所有表名
wb = load_workbook('sample1.xlsx')
print(wb.sheetnames)
# 逐个显示表名
for sheet in wb:
    print(sheet.title)


# 把具体实例表赋值都变量
ws = wb['New Title']
ws.append([11,12,13,14])
ws1 = wb['Sheet']
ws1.append([1,2,3,4])
ws2 = wb['Mysheet1']
ws2.append(['a','b','c','d'])
wb.save('sample2.xlsx')
print(ws)
print(ws1)
print(ws2)

# 遍历表单里单元格值的多种方法：

# 方法1（推荐方式）：一行作为一个元祖展现：
for row in ws.values:
    print(row)
  # 遍历单元格展现（按行遍历）
for row in ws.values:
    for value in row:
        print(value)

# 方法2：用iter按行遍历(iterator:迭代)
for row in ws.iter_rows():
    for cell in row:
        print(cell)  #坐标
        # print(cell.value) #值
  # 按列遍历
for col in ws.iter_cols():
    for cell in col:
        print(cell)
print('------')


# 方法3：用worksheet.max_row,worksheet.max_column按行遍历
for x in range(1,ws.max_row+1):
    for y in range(1,ws.max_column+1):
        print(ws.cell(row=x,column=y))
        print(ws.cell(row=x, column=y).value)
        # from openpyxl.utils import get_column_letter  # 把列值数字转换成字母
        # print(ws[get_column_letter(y)+str(x)])


