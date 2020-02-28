import openpyxl
import sys
from pathlib import Path
excel_file_loaction = Path(sys.argv[0]).parent.joinpath('case2.xlsx')

# # 创建一个工作簿
# wb = openpyxl.Workbook()
# # 创建一个test_case的sheet表单
# wb.create_sheet('test_case')
# # 保存为一个xlsx格式的文件
# wb.save('case2.xlsx')




# # 读取excel中的数据
#第一步：打开工作簿
wb = openpyxl.load_workbook(excel_file_loaction)
# 第二步：选取表单
sh = wb['Sheet']
# 第三步：读取数据
# 参数 row:行  column：列
ce = sh.cell(row=1,column=1)   # 读取第一行，第一列的数据
print(ce.value)
# 按行读取数据 list(sh.rows)
print(list(sh.rows)[0:])     # 按行读取数据，去掉第一行的表头信息数据
for cases in list(sh.rows)[0:]:
    case_id = cases[0].value
    case_excepted = cases[1].value
    case_data = cases[2].value
    print(case_excepted,case_data)


# 关闭工作薄
wb.close()

#写入,写入之前该文件一定要处于关闭状态
wb = openpyxl.load_workbook('case2.xlsx')
sh = wb.active
sh['A6'] = 6
sh['A7'] = 7
sh['A8'] = 8
sh.cell(row=7,column=2,value='Result')
sh.cell(row=8,column=3,value='Amazing')
wb.save('case2.xlsx')

print('Hello Python Excel')





